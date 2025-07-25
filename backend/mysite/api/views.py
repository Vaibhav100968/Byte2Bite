from django.shortcuts import render
from rest_framework import generics, status, serializers
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Customer, InventoryItem, Sale, InventoryReport, Store, Product, Order, OrderItem
from .serializers import CustomerSerializer, InventoryItemSerializer, SaleSerializer, InventoryReportSerializer, StoreSerializer, ProductSerializer, OrderSerializer, OrderItemSerializer
# Stripe import removed
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from accounts.models import User
from accounts.serializers import UserUpdateSerializer
import requests
import base64
import json
import os
from django.conf import settings
from datetime import datetime, timedelta
import pandas as pd
from django.http import FileResponse
import openpyxl
from openpyxl.styles import Font, PatternFill
import logging
from .chatbot.chatbot import analyze_csv, chat_with_gpt

logger = logging.getLogger(__name__)


class CustomerListCreate(generics.ListCreateAPIView):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def analyze_image(request):
    """
    Analyze image using OpenAI GPT-4-Vision API and update inventory
    Accepts base64 encoded image or multipart form data
    """
    try:
        # Get image data from request
        image_data = None

        # Check if image is sent as base64 in JSON
        if request.content_type == 'application/json':
            image_base64 = request.data.get('image')
            if not image_base64:
                return Response(
                    {'error': 'Image data is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            image_data = image_base64

        # Check if image is sent as multipart form data
        elif request.content_type.startswith('multipart/form-data'):
            image_file = request.FILES.get('image')
            if not image_file:
                return Response(
                    {'error': 'Image file is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # Convert file to base64
            image_data = base64.b64encode(image_file.read()).decode('utf-8')

        else:
            return Response(
                {'error': 'Unsupported content type. Use application/json with base64 image or multipart/form-data with image file'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get OpenAI API key from environment variable
        openai_api_key = os.getenv('OPENAI_API_KEY')
        if not openai_api_key:
            return Response(
                {'error': 'OpenAI API key not configured'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Prepare request to OpenAI GPT-4-Vision API
        headers = {
            'Authorization': f'Bearer {openai_api_key}',
            'Content-Type': 'application/json'
        }

        payload = {
            'model': 'gpt-4o-mini',
            'messages': [
                {
                    'role': 'user',
                    'content': [
                        {
                            'type': 'text',
                            'text': 'Count and identify all inventory items in this image. Return JSON with item names and counts only. Format: {"items": {"item_name": count}}'
                        },
                        {
                            'type': 'image_url',
                            'image_url': {
                                'url': f'data:image/jpeg;base64,{image_data}'
                            }
                        }
                    ]
                }
            ],
            'max_tokens': 300
        }

        # Make request to OpenAI API
        response = requests.post(
            'https://api.openai.com/v1/chat/completions',
            headers=headers,
            json=payload,
            timeout=30
        )

        if response.status_code != 200:
            return Response(
                {'error': f'OpenAI API error: {response.text}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Parse OpenAI response
        openai_response = response.json()
        content = openai_response['choices'][0]['message']['content']

        # Try to parse JSON from the response
        try:
            # Extract JSON from the response (in case OpenAI wraps it in markdown)
            import re
            json_match = re.search(r'\{.*\}', content, re.DOTALL)
            if json_match:
                json_str = json_match.group()
                parsed_data = json.loads(json_str)
            else:
                # If no JSON found, return the raw content
                parsed_data = {'raw_response': content}
        except json.JSONDecodeError:
            # If JSON parsing fails, return the raw content
            parsed_data = {'raw_response': content}

        # Update inventory based on detected items
        if 'items' in parsed_data and isinstance(parsed_data['items'], dict):
            updated_items = []
            for item_name, count in parsed_data['items'].items():
                if isinstance(count, (int, float)) and count > 0:
                    # Get or create inventory item
                    item, created = InventoryItem.objects.get_or_create(
                        business=request.user,
                        name=item_name,
                        defaults={
                            'total_added': count,
                            'current_quantity': count
                        }
                    )

                    if not created:
                        # Update existing item
                        item.total_added += count
                        item.current_quantity += count
                        item.save()

                    updated_items.append({
                        'name': item_name,
                        'count': count,
                        'total_added': item.total_added,
                        'current_quantity': item.current_quantity
                    })

            parsed_data['updated_inventory'] = updated_items

        return Response(parsed_data, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error in analyze_image: {str(e)}")
        return Response(
            {'error': f'Internal server error: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class InventoryItemListCreate(generics.ListCreateAPIView):
    serializer_class = InventoryItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return InventoryItem.objects.filter(business=self.request.user)

    def perform_create(self, serializer):
        serializer.save(business=self.request.user)


class InventoryItemDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = InventoryItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return InventoryItem.objects.filter(business=self.request.user)


class SaleListCreate(generics.ListCreateAPIView):
    serializer_class = SaleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Sale.objects.filter(business=self.request.user)

    def perform_create(self, serializer):
        # Get the inventory item
        item = serializer.validated_data['item']
        quantity = serializer.validated_data['quantity']

        # Check if enough inventory is available
        if item.current_quantity < quantity:
            raise serializers.ValidationError(
                f"Insufficient inventory. Available: {item.current_quantity}, Requested: {quantity}"
            )

        # Update inventory
        item.total_sold += quantity
        item.current_quantity -= quantity
        item.save()

        # Create the sale record
        serializer.save(business=self.request.user)


class InventoryReportList(generics.ListAPIView):
    serializer_class = InventoryReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return InventoryReport.objects.filter(business=self.request.user).order_by('-generated_at')


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_inventory_report(request):
    """
    Generate inventory report based on user's reporting frequency
    """
    try:
        user = request.user

        # Create inventory_reports directory if it doesn't exist
        reports_dir = os.path.join(settings.BASE_DIR, 'inventory_reports')
        os.makedirs(reports_dir, exist_ok=True)

        # Get current inventory items
        inventory_items = InventoryItem.objects.filter(business=user)

        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Summary"

        # Add headers
        headers = ['Item Name', 'Total Added',
                   'Total Sold', 'Current Quantity']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Add data
        for row, item in enumerate(inventory_items, 2):
            ws.cell(row=row, column=1, value=item.name)
            ws.cell(row=row, column=2, value=item.total_added)
            ws.cell(row=row, column=3, value=item.total_sold)
            ws.cell(row=row, column=4, value=item.current_quantity)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Generate filename based on frequency
        today = datetime.now()
        if user.reporting_frequency == 'daily':
            filename = f"inventory_daily_{today.strftime('%Y-%m-%d')}.xlsx"
        elif user.reporting_frequency == '3days':
            filename = f"inventory_3days_{today.strftime('%Y-%m-%d')}.xlsx"
        elif user.reporting_frequency == 'weekly':
            week_num = today.isocalendar()[1]
            filename = f"inventory_weekly_W{week_num:02d}.xlsx"
        elif user.reporting_frequency == 'monthly':
            filename = f"inventory_monthly_{today.strftime('%Y-%m')}.xlsx"
        elif user.reporting_frequency == 'custom':
            days = user.custom_reporting_days or 7
            filename = f"inventory_custom_{days}_days_{today.strftime('%Y-%m-%d')}.xlsx"
        else:
            filename = f"inventory_{today.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

        file_path = os.path.join(reports_dir, filename)
        wb.save(file_path)

        # Create report record
        report = InventoryReport.objects.create(
            business=user,
            report_type=user.reporting_frequency,
            file_path=file_path,
            period_start=today.date(),
            period_end=today.date()
        )

        return Response({
            'message': 'Inventory report generated successfully',
            'report_id': report.id,
            'filename': filename
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.error(f"Error generating inventory report: {str(e)}")
        return Response(
            {'error': f'Failed to generate report: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_report(request, report_id):
    """
    Download a specific inventory report
    """
    try:
        report = InventoryReport.objects.get(
            id=report_id, business=request.user)

        if not os.path.exists(report.file_path):
            return Response(
                {'error': 'Report file not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        filename = os.path.basename(report.file_path)
        response = FileResponse(open(report.file_path, 'rb'))
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except InventoryReport.DoesNotExist:
        return Response(
            {'error': 'Report not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Error downloading report: {str(e)}")
        return Response(
            {'error': f'Failed to download report: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_reporting_frequency(request):
    """
    Update user's reporting frequency
    """
    try:
        serializer = UserUpdateSerializer(
            request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'message': 'Reporting frequency updated successfully',
                'reporting_frequency': serializer.data['reporting_frequency'],
                'custom_reporting_days': serializer.data.get('custom_reporting_days')
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f"Error updating reporting frequency: {str(e)}")
        return Response(
            {'error': f'Failed to update reporting frequency: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_test_inventory_data(request):
    """
    Send test inventory data for testing purposes
    """
    try:
        user = request.user

        # Test data with various food items
        test_items = [
            {"name": "Ground Beef", "count": 25},
            {"name": "Chicken Breast", "count": 30},
            {"name": "Pork Chops", "count": 20},
            {"name": "Lettuce", "count": 15},
            {"name": "Tomatoes", "count": 40},
            {"name": "Onions", "count": 35},
            {"name": "Cheese", "count": 18},
            {"name": "Milk", "count": 12},
            {"name": "Bread", "count": 22},
            {"name": "Potatoes", "count": 50},
        ]

        updated_items = []

        # Create or update inventory items
        for item_data in test_items:
            item, created = InventoryItem.objects.get_or_create(
                business=user,
                name=item_data["name"],
                defaults={
                    'total_added': item_data["count"],
                    'current_quantity': item_data["count"]
                }
            )

            if not created:
                # Update existing item
                item.total_added += item_data["count"]
                item.current_quantity += item_data["count"]
                item.save()

            updated_items.append({
                'name': item_data["name"],
                'count': item_data["count"],
                'total_added': item.total_added,
                'current_quantity': item.current_quantity
            })

        # Generate Excel report
        reports_dir = os.path.join(settings.BASE_DIR, 'inventory_reports')
        os.makedirs(reports_dir, exist_ok=True)

        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Inventory Summary"

        # Add headers
        headers = ['Item Name', 'Total Added',
                   'Total Sold', 'Current Quantity']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Add data
        for row, item in enumerate(updated_items, 2):
            ws.cell(row=row, column=1, value=item['name'])
            ws.cell(row=row, column=2, value=item['total_added'])
            ws.cell(row=row, column=3, value=0)  # No sales in test data
            ws.cell(row=row, column=4, value=item['current_quantity'])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Generate filename
        today = datetime.now()
        filename = f"test_inventory_{today.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        file_path = os.path.join(reports_dir, filename)
        wb.save(file_path)

        # Create report record
        report = InventoryReport.objects.create(
            business=user,
            report_type='test',
            file_path=file_path,
            period_start=today.date(),
            period_end=today.date()
        )

        return Response({
            'message': 'Test inventory data sent successfully',
            'updated_items': updated_items,
            'report_id': report.id,
            'filename': filename
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.error(f"Error in send_test_inventory_data: {str(e)}")
        return Response(
            {'error': f'Internal server error: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def chat_with_ai(request):
    """
    Chat with Byte2Bite AI assistant
    """
    try:
        user_message = request.data.get('message')
        csv_data = request.data.get('csv_data')  # CSV data as string or file
        business_profile = request.data.get('business_profile', {})

        # Debug logging
        logger.info(f"Chat API called with message: {user_message}")
        logger.info(f"CSV data provided: {bool(csv_data)}")
        logger.info(f"Business profile provided: {bool(business_profile)}")

        if not user_message:
            return Response(
                {'error': 'Message is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Analyze CSV data if provided
        context_summary = {}
        if csv_data:
            # If CSV data is provided as string, save to temp file
            if isinstance(csv_data, str):
                import tempfile
                with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                    f.write(csv_data)
                    temp_file_path = f.name

                try:
                    context_summary = analyze_csv(temp_file_path)
                    logger.info(f"CSV analysis result: {context_summary}")
                finally:
                    os.unlink(temp_file_path)  # Clean up temp file
            else:
                # Handle file upload
                context_summary = analyze_csv(csv_data)

        # Get AI response with business profile
        response = chat_with_gpt(
            user_message, context_summary, business_profile)
        logger.info(f"AI response generated: {response[:100]}...")

        return Response({
            'response': response,
            'context_summary': context_summary
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error in chat_with_ai: {str(e)}")
        return Response(
            {'error': f'Chat error: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# Store Views
class StoreList(generics.ListAPIView):
    serializer_class = StoreSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Store.objects.filter(is_active=True)


class StoreDetail(generics.RetrieveAPIView):
    serializer_class = StoreSerializer
    permission_classes = [IsAuthenticated]
    queryset = Store.objects.all()


# Product Views
class ProductList(generics.ListAPIView):
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        store_id = self.kwargs.get('store_id')
        return Product.objects.filter(store_id=store_id, in_stock=True)


# Order Views
class OrderListCreate(generics.ListCreateAPIView):
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Order.objects.filter(customer=self.request.user)

    def perform_create(self, serializer):
        # Calculate total amount
        items_data = self.request.data.get('items', [])
        total_amount = 0
        
        for item_data in items_data:
            product = Product.objects.get(id=item_data['product_id'])
            quantity = item_data['quantity']
            total_amount += product.price * quantity

        # Create order
        order = serializer.save(
            customer=self.request.user,
            total_amount=total_amount
        )

        # Create order items
        for item_data in items_data:
            product = Product.objects.get(id=item_data['product_id'])
            OrderItem.objects.create(
                order=order,
                product=product,
                quantity=item_data['quantity'],
                price=product.price
            )

        return order


class OrderDetail(generics.RetrieveAPIView):
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Order.objects.filter(customer=self.request.user)


# Stripe payment views removed - payment integration removed
